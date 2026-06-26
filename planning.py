from io import BytesIO
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import load_workbook


def run_planning(
    input_bytes: bytes,
    target_utilisation_pct: float = 90.0,
    running_thickness_l1: str | None = None,
    running_thickness_l2: str | None = None,
) -> tuple[bytes, dict]:
    """
    3-day planning algorithm with SO-level integrity.

    Line 1:
      - ROOF first
      - then WALL
      - one major roof→wall mode change over the 3-day cycle

    Line 2:
      - ROOF only

    SO selection:
      - Strictly by AGEING descending
      - All-or-nothing at SO level
      - ROOF goes to L2 first if possible, else L1
      - WALL goes to L1 only

    Returns:
      output_bytes, stats
    """

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
        tmp_in.write(input_bytes)
        tmp_in.flush()
        input_path = tmp_in.name

    wb = load_workbook(input_path, data_only=False)
    xls = pd.ExcelFile(input_path)

    demand = pd.read_excel(xls, sheet_name="DEMAND")
    capacity = pd.read_excel(xls, sheet_name="CAPACITY")

    cap_l1 = dict(zip(capacity["TYPE"], capacity["LINE 1 CAPACITY (SQM/HR)"]))
    cap_l2 = dict(zip(capacity["TYPE"], capacity["LINE 2 CAPACITY (SQM/HR)"]))

    day_hours = 24.0
    cycle_days = 3
    cycle_hours = day_hours * cycle_days
    max_hours = cycle_hours * (target_utilisation_pct / 100.0)

    so_ages = demand.groupby("SO NO")["AGEING"].max().to_dict()
    ordered_sos = sorted(so_ages.keys(), key=lambda so: so_ages[so], reverse=True)

    def get_so_rows(so_no):
        sub = demand[demand["SO NO"] == so_no]
        roof_rows = sub[sub["TYPE"].str.contains("ROOF PANEL", na=False)]
        wall_rows = sub[sub["TYPE"].str.contains("WALL PANEL", na=False)]
        return roof_rows, wall_rows

    def runtime_for_rows(rows: pd.DataFrame, line: int) -> float:
        if rows.empty:
            return 0.0
        cap_dict = cap_l1 if line == 1 else cap_l2
        total = 0.0
        for _, r in rows.iterrows():
            cap_val = cap_dict.get(r["TYPE"])
            if cap_val is None or pd.isna(cap_val) or cap_val == 0:
                continue
            total += float(r["TOTAL RMT"]) / float(cap_val)
        return total

    total_runtime_l1 = 0.0
    total_runtime_l2 = 0.0

    line1_roof_rows = []
    line1_wall_rows = []
    line2_roof_rows = []

    cols = [
        "SO NO",
        "W.O NO",
        "TYPE",
        "TOTAL RMT",
        "WIDTH",
        "SQM",
        "TOP COIL",
        "BOTTOM COIL",
        "REMARKS",
        "AGEING",
    ]

    for so in ordered_sos:
        roof_rows, wall_rows = get_so_rows(so)

        has_roof = not roof_rows.empty
        has_wall = not wall_rows.empty

        if not has_roof and not has_wall:
            continue

        rt_roof_l2 = runtime_for_rows(roof_rows, line=2) if has_roof else 0.0
        rt_roof_l1 = runtime_for_rows(roof_rows, line=1) if has_roof else 0.0
        rt_wall_l1 = runtime_for_rows(wall_rows, line=1) if has_wall else 0.0

        if has_roof and has_wall:
            if (
                total_runtime_l2 + rt_roof_l2 <= max_hours
                and total_runtime_l1 + rt_wall_l1 <= max_hours
            ):
                for _, r in wall_rows.iterrows():
                    line1_wall_rows.append({c: r[c] for c in cols})
                for _, r in roof_rows.iterrows():
                    line2_roof_rows.append({c: r[c] for c in cols})
                total_runtime_l1 += rt_wall_l1
                total_runtime_l2 += rt_roof_l2

            elif total_runtime_l1 + rt_roof_l1 + rt_wall_l1 <= max_hours:
                for _, r in roof_rows.iterrows():
                    line1_roof_rows.append({c: r[c] for c in cols})
                for _, r in wall_rows.iterrows():
                    line1_wall_rows.append({c: r[c] for c in cols})
                total_runtime_l1 += (rt_roof_l1 + rt_wall_l1)

        elif has_roof and not has_wall:
            if total_runtime_l2 + rt_roof_l2 <= max_hours:
                for _, r in roof_rows.iterrows():
                    line2_roof_rows.append({c: r[c] for c in cols})
                total_runtime_l2 += rt_roof_l2

            elif total_runtime_l1 + rt_roof_l1 <= max_hours:
                for _, r in roof_rows.iterrows():
                    line1_roof_rows.append({c: r[c] for c in cols})
                total_runtime_l1 += rt_roof_l1

        elif has_wall and not has_roof:
            if total_runtime_l1 + rt_wall_l1 <= max_hours:
                for _, r in wall_rows.iterrows():
                    line1_wall_rows.append({c: r[c] for c in cols})
                total_runtime_l1 += rt_wall_l1

    def thickness_key(type_str: str) -> str:
        if isinstance(type_str, str):
            return type_str.split()[0]
        return ""

    def parse_coil(coil_str: str):
        if not isinstance(coil_str, str):
            return None, ""
        parts = coil_str.split()
        if not parts:
            return None, ""
        try:
            thick = float(parts[0])
        except ValueError:
            thick = None
        color = parts[1] if len(parts) > 1 else ""
        return thick, color

    def sort_with_running(line_rows, running_thickness):
        if not line_rows:
            return []

        def key_fn(d):
            tk = thickness_key(d.get("TYPE"))
            top_thick, top_color = parse_coil(d.get("TOP COIL"))
            bot_thick, bot_color = parse_coil(d.get("BOTTOM COIL"))

            primary = 0 if (running_thickness and tk == running_thickness) else 1
            top_thick_sort = top_thick if top_thick is not None else 1e9
            bot_thick_sort = bot_thick if bot_thick is not None else 1e9
            age = float(d.get("AGEING", 0))

            return (
                primary,
                tk,
                top_color,
                bot_color,
                top_thick_sort,
                bot_thick_sort,
                -age,
            )

        return sorted(line_rows, key=key_fn)

    def count_thickness_changeovers(sorted_rows):
        last_tk = None
        changes = 0
        for d in sorted_rows:
            tk = thickness_key(d.get("TYPE"))
            if last_tk is None:
                last_tk = tk
                continue
            if tk != last_tk:
                changes += 1
                last_tk = tk
        return changes

    l1_roof_sorted = sort_with_running(line1_roof_rows, running_thickness_l1)
    l1_wall_sorted = sort_with_running(line1_wall_rows, running_thickness_l1)
    line1_rows_sorted = l1_roof_sorted + l1_wall_sorted

    line2_rows_sorted = sort_with_running(line2_roof_rows, running_thickness_l2)

    def assign_days(sorted_rows, line):
        out = []
        cumulative = 0.0
        cap_dict = cap_l1 if line == 1 else cap_l2

        for d in sorted_rows:
            cap_val = cap_dict.get(d["TYPE"])
            runtime = 0.0
            if cap_val is not None and not pd.isna(cap_val) and cap_val != 0:
                runtime = float(d["TOTAL RMT"]) / float(cap_val)

            start_hour = cumulative
            if start_hour < 24:
                day = 1
            elif start_hour < 48:
                day = 2
            else:
                day = 3

            row_copy = dict(d)
            row_copy["RUN TIME"] = runtime
            row_copy["DAY 1"] = d["TOTAL RMT"] if day == 1 else None
            row_copy["DAY 2"] = d["TOTAL RMT"] if day == 2 else None
            row_copy["DAY 3"] = d["TOTAL RMT"] if day == 3 else None
            row_copy["DAY"] = day

            out.append(row_copy)
            cumulative += runtime

        return out

    line1_rows_with_days = assign_days(line1_rows_sorted, line=1)
    line2_rows_with_days = assign_days(line2_rows_sorted, line=2)

    def build_day_stats(rows_with_days):
        out = {
            "day1_runtime": 0.0,
            "day2_runtime": 0.0,
            "day3_runtime": 0.0,
            "day1_sqm": 0.0,
            "day2_sqm": 0.0,
            "day3_sqm": 0.0,
            "day1_rows": 0,
            "day2_rows": 0,
            "day3_rows": 0,
        }

        for d in rows_with_days:
            rt = float(d.get("RUN TIME", 0) or 0)
            sqm = float(d.get("SQM", 0) or 0)
            day = d.get("DAY")

            if day == 1:
                out["day1_runtime"] += rt
                out["day1_sqm"] += sqm
                out["day1_rows"] += 1
            elif day == 2:
                out["day2_runtime"] += rt
                out["day2_sqm"] += sqm
                out["day2_rows"] += 1
            elif day == 3:
                out["day3_runtime"] += rt
                out["day3_sqm"] += sqm
                out["day3_rows"] += 1

        return out

    day_stats_l1 = build_day_stats(line1_rows_with_days)
    day_stats_l2 = build_day_stats(line2_rows_with_days)

    total_sqm_l1 = sum(float(d.get("SQM", 0) or 0) for d in line1_rows_sorted)
    total_sqm_l2 = sum(float(d.get("SQM", 0) or 0) for d in line2_rows_sorted)

    changeovers_l1 = count_thickness_changeovers(line1_rows_sorted)
    changeovers_l2 = count_thickness_changeovers(line2_rows_sorted)

    util_l1_pct = (total_runtime_l1 / cycle_hours) * 100.0 if cycle_hours else 0.0
    util_l2_pct = (total_runtime_l2 / cycle_hours) * 100.0 if cycle_hours else 0.0

    ws1 = wb["LINE 1 SCHEDULE"]
    ws2 = wb["LINE 2 SCHEDULE"]

    header_row = 3
    data_start_row = header_row + 1
    max_data_row = 200

    for ws in (ws1, ws2):
        for r in range(data_start_row, max_data_row + 1):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None

    ws1["L3"] = "DAY 1"
    ws1["M3"] = "DAY 2"
    ws1["N3"] = "DAY 3"

    ws2["L3"] = "DAY 1"
    ws2["M3"] = "DAY 2"
    ws2["N3"] = "DAY 3"

    target_cols = cols

    for idx, row_data in enumerate(line1_rows_with_days):
        excel_row = data_start_row + idx
        if excel_row > max_data_row:
            break

        for col_idx, key in enumerate(target_cols, start=1):
            ws1.cell(row=excel_row, column=col_idx).value = row_data.get(key)

        ws1.cell(row=excel_row, column=11).value = row_data.get("RUN TIME")
        ws1.cell(row=excel_row, column=12).value = row_data.get("DAY 1")
        ws1.cell(row=excel_row, column=13).value = row_data.get("DAY 2")
        ws1.cell(row=excel_row, column=14).value = row_data.get("DAY 3")

    for idx, row_data in enumerate(line2_rows_with_days):
        excel_row = data_start_row + idx
        if excel_row > max_data_row:
            break

        for col_idx, key in enumerate(target_cols, start=1):
            ws2.cell(row=excel_row, column=col_idx).value = row_data.get(key)

        ws2.cell(row=excel_row, column=11).value = row_data.get("RUN TIME")
        ws2.cell(row=excel_row, column=12).value = row_data.get("DAY 1")
        ws2.cell(row=excel_row, column=13).value = row_data.get("DAY 2")
        ws2.cell(row=excel_row, column=14).value = row_data.get("DAY 3")

    for ws, total_runtime, util_pct, total_sqm in [
        (ws1, total_runtime_l1, util_l1_pct, total_sqm_l1),
        (ws2, total_runtime_l2, util_l2_pct, total_sqm_l2),
    ]:
        ws["K1"] = util_pct / 100.0
        ws["K2"] = total_runtime
        ws["L2"] = cycle_hours
        ws["M2"] = total_sqm

    out_bytes = BytesIO()
    wb.save(out_bytes)
    out_bytes.seek(0)
    output_bytes = out_bytes.read()

    stats = {
        "total_runtime_l1": float(total_runtime_l1),
        "total_runtime_l2": float(total_runtime_l2),
        "util_l1_pct": float(util_l1_pct),
        "util_l2_pct": float(util_l2_pct),
        "total_sqm_l1": float(total_sqm_l1),
        "total_sqm_l2": float(total_sqm_l2),
        "thickness_changeovers_l1": int(changeovers_l1),
        "thickness_changeovers_l2": int(changeovers_l2),
        "rows_l1": len(line1_rows_sorted),
        "rows_l2": len(line2_rows_sorted),
        "planning_horizon_hours": cycle_hours,
        "planning_days": cycle_days,
        "day_stats_l1": day_stats_l1,
        "day_stats_l2": day_stats_l2,
    }

    return output_bytes, stats
